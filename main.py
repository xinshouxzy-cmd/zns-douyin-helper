# -*- coding: utf-8 -*-
"""
遵农商·抖音客服助手
多账号抖音私信+评论自动回复工具
基于 PyQt5 + Selenium
"""

import os, sys, json, csv, traceback, html as _html
from datetime import datetime
from threading import Event

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QPushButton, QLabel, QTextEdit, QLineEdit, QTabBar,
    QCheckBox, QGroupBox, QScrollArea, QMessageBox, QFileDialog, QFrame,
    QInputDialog, QListWidget, QListWidgetItem, QStackedWidget, QSizePolicy,
    QGraphicsDropShadowEffect, QDialog, QFormLayout, QDialogButtonBox
)
from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal, QSize
from PyQt5.QtGui import QFont, QColor, QPalette, QTextCursor, QIcon, QPixmap, QPainter

from worker import AccountWorker, BASE_DIR
from calibration_data import (
    CALIBRATION_STEPS, get_calibration_status, load_calibration,
    save_shared_calibration, save_account_calibration, copy_shared_to_account
)

try:
    from _version import VERSION
except Exception:
    VERSION = "v2.0"

# ── 配置 ──────────────────────────────────────────
APP_TITLE = f"遵农商·抖音客服助手 {VERSION} — 辛振宇"
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
DEFAULT_PM_REPLY = "遵义地区政策了解，留下☎️"
DEFAULT_CMT_REPLY = "具体抖音✉️"
PM_POLL = 5
CMT_POLL = 30


def load_config():
    if not os.path.exists(CONFIG_FILE):
        return {"accounts": []}
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# ── 微信风格配色 ───────────────────────────────────
C_SIDEBAR_BG = "#2C2C2C"
C_SIDEBAR_HOVER = "#3A3A3A"
C_SIDEBAR_ACTIVE = "#3A3A3A"
C_MAIN_BG = "#F0F0F0"
C_CARD_BG = "#FFFFFF"
C_TEXT_PRIMARY = "#1A1A1A"
C_TEXT_SECONDARY = "#888888"
C_TEXT_SIDEBAR = "#CCCCCC"
C_TEXT_SIDEBAR_ACTIVE = "#FFFFFF"
C_ACCENT = "#07C160"
C_ACCENT_HOVER = "#06AD56"
C_RED = "#FA5151"
C_BORDER = "#E5E5E5"
C_STATUS_RUNNING = "#07C160"
C_STATUS_STOPPED = "#B0B0B0"
C_LOG_BG = "#F8F8F8"
C_BTN_DISABLED = "#C0C0C0"

# 兼容旧引用
C_GREEN = C_ACCENT
C_YELLOW = "#E6A23C"
C_TEXT = C_TEXT_PRIMARY
C_BG = C_MAIN_BG
C_PANEL = C_CARD_BG
C_INPUT = C_CARD_BG

STYLE = f"""
QMainWindow {{ background: {C_MAIN_BG}; }}
QWidget {{
    font-size: 14px;
    font-family: "PingFang SC", "Microsoft YaHei", "SF Pro Display", sans-serif;
}}
QLineEdit {{
    background: {C_CARD_BG}; color: {C_TEXT_PRIMARY};
    border: 1px solid {C_BORDER}; border-radius: 6px;
    padding: 10px 14px; font-size: 14px;
}}
QLineEdit:focus {{ border-color: {C_ACCENT}; background: #F0FFF5; }}
QLineEdit:disabled {{ background: #F5F5F5; color: #BBB; }}
QTextEdit {{
    background: {C_LOG_BG}; color: {C_TEXT_SECONDARY};
    border: 1px solid {C_BORDER}; border-radius: 6px;
    padding: 8px 12px; font-size: 12px;
    font-family: "SF Mono", "Menlo", "Consolas", "Courier New", monospace;
}}
QScrollArea {{ border: none; background: transparent; }}
QCheckBox {{
    color: {C_TEXT_PRIMARY}; spacing: 8px; font-size: 14px;
}}
QCheckBox::indicator {{
    width: 20px; height: 20px;
    border: 2px solid {C_BORDER}; border-radius: 4px;
    background: {C_CARD_BG};
}}
QCheckBox::indicator:checked {{
    background: {C_ACCENT}; border-color: {C_ACCENT};
}}
QLabel {{ color: {C_TEXT_PRIMARY}; }}
QScrollBar:vertical {{
    background: transparent; width: 6px; margin: 0;
}}
QScrollBar::handle:vertical {{
    background: #CCC; border-radius: 3px; min-height: 30px;
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QScrollBar:horizontal {{
    background: transparent; height: 6px;
}}
QScrollBar::handle:horizontal {{
    background: #CCC; border-radius: 3px; min-width: 30px;
}}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{ width: 0; }}
"""


def _btn_primary():
    """微信绿主按钮"""
    return f"""
        QPushButton {{
            background: {C_ACCENT}; color: white; border: none;
            border-radius: 6px; padding: 10px 24px;
            font-size: 14px; font-weight: bold;
        }}
        QPushButton:hover {{ background: {C_ACCENT_HOVER}; }}
        QPushButton:pressed {{ background: #05944A; }}
        QPushButton:disabled {{ background: {C_BTN_DISABLED}; color: #999; }}
    """


def _btn_danger():
    """红色按钮（停止）"""
    return f"""
        QPushButton {{
            background: {C_RED}; color: white; border: none;
            border-radius: 6px; padding: 10px 24px;
            font-size: 14px; font-weight: bold;
        }}
        QPushButton:hover {{ background: #E04848; }}
        QPushButton:pressed {{ background: #C73E3E; }}
        QPushButton:disabled {{ background: {C_BTN_DISABLED}; color: #999; }}
    """


def _btn_default():
    """灰色次要按钮"""
    return f"""
        QPushButton {{
            background: #E5E5E5; color: {C_TEXT_PRIMARY}; border: none;
            border-radius: 6px; padding: 10px 24px;
            font-size: 14px;
        }}
        QPushButton:hover {{ background: #D5D5D5; }}
        QPushButton:pressed {{ background: #C5C5C5; }}
        QPushButton:disabled {{ background: #F0F0F0; color: #BBB; }}
    """


# 兼容旧 _btn() 调用
def _btn(color, text_color="white"):
    return f"""
        QPushButton {{
            background: {color}; color: {text_color}; border: none;
            border-radius: 6px; padding: 10px 24px;
            font-size: 14px; font-weight: bold;
        }}
        QPushButton:hover {{ opacity: 0.85; }}
        QPushButton:pressed {{ background: #333; }}
        QPushButton:disabled {{ background: #555; color: #888; }}
    """


class Card(QFrame):
    """白色圆角卡片"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        self.setStyleSheet(f"""
            #card {{
                background: {C_CARD_BG};
                border: 1px solid {C_BORDER};
                border-radius: 10px;
            }}
        """)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)


class SidebarItem(QWidget):
    """侧边栏账号项"""
    clicked = pyqtSignal(int)
    remove_requested = pyqtSignal(int)

    def __init__(self, idx, name, parent=None):
        super().__init__(parent)
        self.idx = idx
        self._name = name
        self._active = False
        self._is_running = False
        self.setFixedHeight(64)
        self.setCursor(Qt.PointingHandCursor)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(12, 8, 12, 8)
        lay.setSpacing(10)

        self.dot = QLabel()
        self.dot.setFixedSize(10, 10)
        self._update_dot()
        lay.addWidget(self.dot)

        tv = QVBoxLayout()
        tv.setSpacing(2)
        self.lbl_name = QLabel(self._name)
        self.lbl_name.setStyleSheet(f"color:{C_TEXT_SIDEBAR};font-size:14px;font-weight:500;")
        tv.addWidget(self.lbl_name)
        self.lbl_status = QLabel("已停止")
        self.lbl_status.setStyleSheet("color:#888;font-size:11px;")
        tv.addWidget(self.lbl_status)
        lay.addLayout(tv, 1)

        self.btn_close = QPushButton("×")
        self.btn_close.setFixedSize(20, 20)
        self.btn_close.setStyleSheet(f"""
            QPushButton {{ background:transparent;color:#888;border:none;font-size:16px;font-weight:bold;padding:0; }}
            QPushButton:hover {{ color:{C_RED};background:rgba(255,255,255,0.1);border-radius:10px; }}
        """)
        self.btn_close.clicked.connect(lambda: self.remove_requested.emit(self.idx))
        self.btn_close.setVisible(False)
        lay.addWidget(self.btn_close)

    def _update_dot(self):
        color = C_STATUS_RUNNING if self._is_running else C_STATUS_STOPPED
        r = 5
        pix = QPixmap(r * 2 + 2, r * 2 + 2)
        pix.fill(Qt.transparent)
        p = QPainter(pix)
        p.setRenderHint(QPainter.Antialiasing)
        p.setBrush(QColor(color))
        p.setPen(Qt.NoPen)
        p.drawEllipse(1, 1, r * 2, r * 2)
        p.end()
        self.dot.setPixmap(pix)

    def set_status(self, is_running, text):
        self._is_running = is_running
        self._update_dot()
        self.lbl_status.setText(text)
        if is_running:
            self.lbl_status.setStyleSheet(f"color:{C_STATUS_RUNNING};font-size:11px;")
        else:
            self.lbl_status.setStyleSheet("color:#888;font-size:11px;")

    def set_name(self, name):
        self._name = name
        self.lbl_name.setText(name)

    def set_active(self, active):
        self._active = active
        if active:
            self.setStyleSheet(f"""
                SidebarItem {{
                    background: {C_SIDEBAR_ACTIVE};
                    border-left: 3px solid {C_ACCENT};
                }}
            """)
            self.lbl_name.setStyleSheet(f"color:{C_TEXT_SIDEBAR_ACTIVE};font-size:14px;font-weight:500;")
        else:
            self.setStyleSheet("SidebarItem { border-left: 3px solid transparent; }")
            self.lbl_name.setStyleSheet(f"color:{C_TEXT_SIDEBAR};font-size:14px;font-weight:400;")

    def enterEvent(self, event):
        if not self._active:
            self.setStyleSheet(f"SidebarItem {{ background: {C_SIDEBAR_HOVER}; border-left: 3px solid transparent; }}")
        self.btn_close.setVisible(True)
        super().enterEvent(event)

    def leaveEvent(self, event):
        if not self._active:
            self.setStyleSheet("SidebarItem { border-left: 3px solid transparent; }")
        self.btn_close.setVisible(False)
        super().leaveEvent(event)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.clicked.emit(self.idx)
        super().mousePressEvent(event)


# ── 手动校准向导对话框 ───────────────────────────────
class CalibrationWizard(QDialog):
    """3步手动校准向导 - 浏览器内连点5次同一位置自动确认"""

    def __init__(self, worker, account_name, parent=None):
        super().__init__(parent)
        self.worker = worker
        self.account_name = account_name
        self.current_step = 0  # 0=未开始, 1-3=步骤
        self.captured = {}  # {step_id: {x, y}}
        self._cancelled = False

        self.setWindowTitle(f"📐 评论坐标校准 - {account_name}")
        self.setMinimumSize(480, 350)
        self.setModal(True)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        self._build_ui()
        self._update_step_display()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(14)

        # 标题
        title = QLabel("📐 校准评论回复坐标（3步）")
        title.setStyleSheet("font-size:18px; font-weight:bold; color:#1A1A1A;")
        lay.addWidget(title)

        desc = QLabel("在浏览器中连续点击同一位置5次即可自动记录。\n同一台电脑只需校准一次，其他账号自动共享。")
        desc.setStyleSheet("color:#666; font-size:13px;")
        desc.setWordWrap(True)
        lay.addWidget(desc)

        # 进度条
        self.pb_frame = QFrame()
        self.pb_frame.setFixedHeight(8)
        self.pb_frame.setStyleSheet(f"background:#E5E5E5; border-radius:4px;")
        self.pb_lay = QHBoxLayout(self.pb_frame)
        self.pb_lay.setContentsMargins(0, 0, 0, 0)
        self.pb_bar = QFrame()
        self.pb_bar.setStyleSheet(f"background:{C_ACCENT}; border-radius:4px;")
        self.pb_lay.addWidget(self.pb_bar)
        self.pb_lay.addStretch()
        lay.addWidget(self.pb_frame)

        # 当前步骤标题
        self.lbl_step_title = QLabel("")
        self.lbl_step_title.setStyleSheet("font-size:16px; font-weight:bold; color:#1A1A1A;")
        self.lbl_step_title.setWordWrap(True)
        lay.addWidget(self.lbl_step_title)

        # 描述
        self.lbl_desc = QLabel("")
        self.lbl_desc.setStyleSheet("color:#444; font-size:14px;")
        self.lbl_desc.setWordWrap(True)
        lay.addWidget(self.lbl_desc)

        # 提示框
        self.lbl_tip = QLabel("")
        self.lbl_tip.setStyleSheet(f"""
            background: #FFF8E1; color: #795548;
            padding: 10px 14px; border-radius: 6px;
            font-size: 13px; border: 1px solid #FFE082;
        """)
        self.lbl_tip.setWordWrap(True)
        self.lbl_tip.setVisible(False)
        lay.addWidget(self.lbl_tip)

        # 状态
        self.lbl_status = QLabel("")
        self.lbl_status.setStyleSheet("color:#07C160; font-size:13px; font-weight:bold;")
        self.lbl_status.setWordWrap(True)
        lay.addWidget(self.lbl_status)

        lay.addStretch()

        # 按钮
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)
        btn_row.addStretch()

        self.btn_start = QPushButton("▶ 开始校准")
        self.btn_start.setStyleSheet(_btn_primary())
        self.btn_start.setFixedHeight(38)
        self.btn_start.clicked.connect(self._start)
        btn_row.addWidget(self.btn_start)

        self.btn_cancel = QPushButton("取消")
        self.btn_cancel.setStyleSheet(_btn_default())
        self.btn_cancel.setFixedHeight(38)
        self.btn_cancel.clicked.connect(self._cancel)
        btn_row.addWidget(self.btn_cancel)

        lay.addLayout(btn_row)

    def _update_step_display(self):
        s = self.current_step
        if s == 0:
            self.lbl_step_title.setText("准备校准")
            self.lbl_desc.setText("点击「开始校准」后，在浏览器中依次连续点击3个关键位置，每个位置连点5次即可自动确认。")
            self.lbl_status.setText("💡 请确保浏览器已显示抖音首页")
            self.lbl_tip.setVisible(False)
            self.btn_start.setVisible(True)
            self._update_progress(0)
        elif 1 <= s <= 3:
            info = CALIBRATION_STEPS[s - 1]
            done_str = " · ".join([
                f"{'✅' if CALIBRATION_STEPS[i]['id'] in self.captured else '⬜'} {CALIBRATION_STEPS[i]['label']}"
                for i in range(3)
            ])
            self.lbl_step_title.setText(f"第 {s}/3 步：{info['label']}")
            self.lbl_desc.setText(info['desc'])
            self.lbl_tip.setText(f"状态：{done_str}")
            self.lbl_tip.setVisible(True)
            self.btn_start.setVisible(False)
            self._update_progress(s - 1)

    def _update_progress(self, step):
        pct = step / 3
        self.pb_bar.setFixedWidth(int(pct * self.pb_frame.width()))

    def _start(self):
        if not self.worker._d:
            self.lbl_status.setText("❌ 浏览器未就绪，请先启动账号")
            return

        # 进入校准模式
        self.worker.enter_calibration_mode(as_shared=True)

        # 连接信号
        self.worker.calib_step.connect(self._on_calib_step_signal)
        self.worker.calib_captured.connect(self._on_captured_signal)

        # 依次执行3步
        for step in range(1, 4):
            if self._cancelled:
                break

            self.current_step = step
            self._update_step_display()
            self.lbl_status.setText(f"⏳ 第{step}/3步：请在浏览器中连点5次...")
            QApplication.processEvents()

            # 调用 worker 捕获坐标（阻塞，等待5连点或超时）
            result = self.worker.do_calibration_step(step)

            if result:
                self.captured[result["step_id"]] = {"x": result["x"], "y": result["y"]}
                self.lbl_status.setText(f"✅ 第{step}/3步已确认: ({result['x']}, {result['y']})")
                QApplication.processEvents()
            else:
                self.lbl_status.setText(f"⚠ 第{step}/3步超时，已跳过")
                QApplication.processEvents()

        self._finish()

    def _on_calib_step_signal(self, name, step, text):
        """来自 worker 的步骤信号"""
        pass  # 同步处理

    def _on_captured_signal(self, name, step, x, y):
        """来自 worker 的捕获信号"""
        pass  # 同步处理

    def _cancel(self):
        self._cancelled = True

        # 断开信号
        try:
            self.worker.calib_step.disconnect(self._on_calib_step_signal)
            self.worker.calib_captured.disconnect(self._on_captured_signal)
        except:
            pass

        # 退出校准模式（不保存）
        self.worker.exit_calibration_mode(None)
        self.reject()

    def _finish(self):
        self.lbl_status.setText("💾 正在保存...")
        self.btn_cancel.setVisible(False)
        QApplication.processEvents()

        # 断开信号
        try:
            self.worker.calib_step.disconnect(self._on_calib_step_signal)
            self.worker.calib_captured.disconnect(self._on_captured_signal)
        except:
            pass

        # 保存
        ok = self.worker.exit_calibration_mode(self.captured)
        if ok:
            self.lbl_step_title.setText("🎉 校准完成！")
            self.lbl_desc.setText(f"已保存 {len(self.captured)}/3 个步骤的坐标。\n本机所有账号均可使用此校准数据。")
            self.lbl_tip.setVisible(False)
            self._update_progress(3)
        else:
            self.lbl_step_title.setText("⚠ 校准不完整")
            self.lbl_desc.setText(f"仅捕获 {len(self.captured)}/3 个步骤，至少需要3步。\n请重新校准。")

        btn_close = QPushButton("关闭")
        btn_close.setStyleSheet(_btn_primary())
        btn_close.setFixedHeight(38)
        btn_close.clicked.connect(self.accept)
        # 找到按钮布局并添加
        for i in range(self.layout().count()):
            item = self.layout().itemAt(i)
            if isinstance(item, QHBoxLayout) and item.itemAt(0) and isinstance(item.itemAt(0).widget(), QWidget):
                # 在 stretch 之前插入
                pass

        QMessageBox.information(self, "校准完成",
            f"✅ 评论坐标校准完成！\n\n"
            f"已保存 {len(self.captured)}/7 个步骤。\n"
            f"本机所有账号均可共享使用。\n\n"
            f"现在可以点击「确认已登录」开始自动运行。")

    def _cancel(self):
        self._cancelled = True
        try:
            self.worker.calib_step.disconnect(self._on_calib_step_signal)
            self.worker.calib_captured.disconnect(self._on_captured_signal)
        except:
            pass
        self.worker._calibration_mode = False
        self.worker._calib_event.set()
        self.reject()


# ── 每个账号的页面 ─────────────────────────────────
class AccountPage(QWidget):
    def __init__(self, idx, cfg, main_win):
        super().__init__()
        self.idx = idx
        self.cfg = cfg
        self.main = main_win
        self.worker = None
        self._pm_count = 0
        self._cmt_count = 0
        self._in_login_wait = False
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(16)

        # ── 标题行：名称 + 状态徽章 ──
        title_row = QHBoxLayout()
        lbl_title = QLabel("⚙️ 账号设置")
        lbl_title.setStyleSheet(f"font-size:20px; font-weight:bold; color:{C_TEXT_PRIMARY};")
        title_row.addWidget(lbl_title)
        title_row.addStretch()
        self.lb_status = QLabel("⏸ 未启动")
        self.lb_status.setStyleSheet(f"""
            background: #F0F0F0; color: {C_TEXT_SECONDARY};
            padding: 4px 14px; border-radius: 12px; font-size: 13px;
        """)
        title_row.addWidget(self.lb_status)
        lay.addLayout(title_row)

        # ── 账号名称卡片 ──
        card_name = Card()
        cn_lay = QVBoxLayout(card_name)
        cn_lay.setContentsMargins(20, 16, 20, 16)
        cn_lay.setSpacing(8)
        lbl_n = QLabel("🏷 账号名称")
        lbl_n.setStyleSheet(f"font-weight:bold; color:{C_TEXT_PRIMARY}; font-size:14px;")
        cn_lay.addWidget(lbl_n)
        self.le_name = QLineEdit(self.cfg.get("name", ""))
        self.le_name.setPlaceholderText("输入账号名称（侧边栏将自动更新）")
        self.le_name.textChanged.connect(self._on_name_changed)
        cn_lay.addWidget(self.le_name)
        lay.addWidget(card_name)

        # ── 确认登录 + 手动校准 ──
        self.btn_login_row = QWidget()
        blr_lay = QHBoxLayout(self.btn_login_row)
        blr_lay.setContentsMargins(0, 0, 0, 0)
        blr_lay.setSpacing(8)

        self.btn_calibrate = QPushButton("📐 手动校准评论")
        self.btn_calibrate.setStyleSheet(_btn_default())
        self.btn_calibrate.clicked.connect(self._open_calibration_wizard)
        self.btn_calibrate.setVisible(False)
        self.btn_calibrate.setFixedHeight(42)
        self.btn_calibrate.setToolTip("登录前手动标记7个评论操作位置，同电脑只需校准一次")
        blr_lay.addWidget(self.btn_calibrate)

        self.btn_login = QPushButton("✓ 确认已扫码登录")
        self.btn_login.setStyleSheet(_btn_primary())
        self.btn_login.clicked.connect(self._confirm_login)
        self.btn_login.setVisible(False)
        self.btn_login.setFixedHeight(42)
        blr_lay.addWidget(self.btn_login)

        self.btn_login_row.setVisible(False)
        lay.addWidget(self.btn_login_row)

        # ── 私信回复卡片 ──
        card_pm = Card()
        cp_lay = QVBoxLayout(card_pm)
        cp_lay.setContentsMargins(20, 16, 20, 16)
        cp_lay.setSpacing(10)
        pm_hdr = QHBoxLayout()
        lbl_pm = QLabel("💬 私信自动回复")
        lbl_pm.setStyleSheet(f"font-weight:bold; color:{C_TEXT_PRIMARY}; font-size:14px;")
        pm_hdr.addWidget(lbl_pm)
        pm_hdr.addStretch()
        self.cb_pm = QCheckBox("启用")
        self.cb_pm.setChecked(self.cfg.get("pm_enabled", True))
        self.cb_pm.toggled.connect(self._save)
        pm_hdr.addWidget(self.cb_pm)
        cp_lay.addLayout(pm_hdr)
        lbl_tip = QLabel("回复话术：")
        lbl_tip.setStyleSheet(f"color:{C_TEXT_SECONDARY}; font-size:12px;")
        cp_lay.addWidget(lbl_tip)
        self.le_pm = QLineEdit(self.cfg.get("pm_reply", DEFAULT_PM_REPLY))
        self.le_pm.setPlaceholderText("私信回复话术...")
        self.le_pm.textChanged.connect(self._save)
        cp_lay.addWidget(self.le_pm)
        lay.addWidget(card_pm)

        # ── 评论回复卡片 ──
        card_cmt = Card()
        cc_lay = QVBoxLayout(card_cmt)
        cc_lay.setContentsMargins(20, 16, 20, 16)
        cc_lay.setSpacing(10)
        cmt_hdr = QHBoxLayout()
        lbl_cmt = QLabel("📝 评论自动回复")
        lbl_cmt.setStyleSheet(f"font-weight:bold; color:{C_TEXT_PRIMARY}; font-size:14px;")
        cmt_hdr.addWidget(lbl_cmt)
        cmt_hdr.addStretch()
        self.cb_cmt = QCheckBox("启用")
        self.cb_cmt.setChecked(self.cfg.get("comment_enabled", True))
        self.cb_cmt.toggled.connect(self._save)
        cmt_hdr.addWidget(self.cb_cmt)
        cc_lay.addLayout(cmt_hdr)
        lbl_tip2 = QLabel("回复话术：")
        lbl_tip2.setStyleSheet(f"color:{C_TEXT_SECONDARY}; font-size:12px;")
        cc_lay.addWidget(lbl_tip2)
        self.le_cmt = QLineEdit(self.cfg.get("comment_reply", DEFAULT_CMT_REPLY))
        self.le_cmt.setPlaceholderText("评论回复话术...")
        self.le_cmt.textChanged.connect(self._save)
        cc_lay.addWidget(self.le_cmt)
        lay.addWidget(card_cmt)

        lay.addStretch()

        # ── 操作按钮 ──
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)
        self.btn_start = QPushButton("▶ 启动")
        self.btn_start.setStyleSheet(_btn_primary())
        self.btn_start.setFixedHeight(42)
        self.btn_start.clicked.connect(self._toggle)
        self.btn_export = QPushButton("📊 导出数据")
        self.btn_export.setStyleSheet(_btn_default())
        self.btn_export.setFixedHeight(42)
        self.btn_export.clicked.connect(self._export_one)
        btn_row.addStretch()
        btn_row.addWidget(self.btn_start)
        btn_row.addWidget(self.btn_export)
        lay.addLayout(btn_row)

    def _on_name_changed(self, txt):
        self._save()
        self.main._update_sidebar_name(self.idx, txt.strip() or f"账号{self.idx+1}")

    def _save(self):
        self.cfg["name"] = self.le_name.text().strip() or f"账号{self.idx+1}"
        self.cfg["pm_enabled"] = self.cb_pm.isChecked()
        self.cfg["pm_reply"] = self.le_pm.text()
        self.cfg["comment_enabled"] = self.cb_cmt.isChecked()
        self.cfg["comment_reply"] = self.le_cmt.text()
        cfg = load_config()
        if self.idx < len(cfg.get("accounts", [])):
            cfg["accounts"][self.idx] = self.cfg
            save_config(cfg)

    def _set_status_ui(self, text, color, is_running=False, status_text=""):
        self.lb_status.setText(text)
        self.lb_status.setStyleSheet(f"""
            background: {color}20; color: {color};
            padding: 4px 14px; border-radius: 12px;
            font-size: 13px; font-weight: bold;
        """)
        if status_text:
            self.main._update_sidebar_status(self.idx, is_running, status_text)

    def _confirm_login(self):
        """用户点击「确认已登录」"""
        if self.worker:
            self.worker.confirm_login()
            self.btn_login_row.setVisible(False)
            self._in_login_wait = False
            self._set_status_ui("登录确认中...", C_ACCENT, True, "登录确认中...")

    def _open_calibration_wizard(self):
        """打开手动校准向导"""
        if not self.worker or not self.worker._d:
            QMessageBox.warning(self, "无法校准", "浏览器尚未就绪，请先启动此账号。")
            return

        # 检查是否已经登录（登录后不允许校准）
        if not self._in_login_wait:
            QMessageBox.warning(self, "无法校准",
                "程序已开始自动运行，无法进行校准。\n\n"
                "如需校准，请先停止并重新启动此账号，\n"
                "在扫码阶段点击「📐 手动校准」。")
            return

        wizard = CalibrationWizard(self.worker, self.cfg.get("name", "?"), self)
        wizard.exec_()

        # 校准完成后，如果有数据，标记为已校准
        status = get_calibration_status(self.cfg.get("name", ""))
        if status["has_shared"] or status["has_account"]:
            self._set_status_ui("✅ 校准完成，请确认登录", C_GREEN, True, "已校准")

    def _toggle(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.btn_start.setText("⏳ 停止中...")
            self.btn_start.setEnabled(False)
        else:
            self._save()
            self.worker = AccountWorker(self.cfg, PM_POLL, CMT_POLL)
            self.worker.log.connect(self.main._append_log)
            self.worker.status.connect(self._on_status)
            self.worker.waiting_login.connect(self._on_waiting_login)
            self.worker.pm_cnt.connect(self._on_pm_cnt)
            self.worker.cmt_cnt.connect(self._on_cmt_cnt)
            self.worker.stopped.connect(self._on_stopped)
            self.worker.start()
            self.btn_start.setText("⏹ 停止")
            self.btn_start.setStyleSheet(_btn_danger())
            self._set_status_ui("启动中...", C_ACCENT, True, "启动中...")

    def _on_waiting_login(self, name):
        if name == self.cfg.get("name"):
            self._in_login_wait = True
            self.btn_login_row.setVisible(True)
            self.btn_login.setVisible(True)
            self.btn_calibrate.setVisible(True)
            self._set_status_ui("📱 请扫码登录", C_YELLOW, True, "等待扫码登录")

    def _on_status(self, name, s):
        if name == self.cfg.get("name"):
            self._set_status_ui(s, C_ACCENT, True, s)

    def _on_pm_cnt(self, name, n):
        if name == self.cfg.get("name"):
            self._pm_count = n

    def _on_cmt_cnt(self, name, n):
        if name == self.cfg.get("name"):
            self._cmt_count = n

    def _on_stopped(self, name):
        if name == self.cfg.get("name"):
            self.worker = None
            self.btn_start.setText("▶ 启动")
            self.btn_start.setStyleSheet(_btn_primary())
            self.btn_start.setEnabled(True)
            self.btn_login_row.setVisible(False)
            self._in_login_wait = False
            self._set_status_ui("⏸ 已停止", C_TEXT_SECONDARY, False, "已停止")

    def _export_one(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from worker import load_replied

        f, _ = QFileDialog.getSaveFileName(self, "导出数据",
            f"{self.cfg.get('name','账号')}_{datetime.now().strftime('%m%d')}.xlsx",
            "Excel (*.xlsx)")
        if not f:
            return

        # ── 暂停 worker（参照 v42.1：导出时停止监控，完成后恢复）──
        was_running = self.worker and self.worker.isRunning()
        worker_ref = self.worker
        if was_running:
            # 断开 stopped 信号，防止 _on_stopped 干扰恢复流程
            try:
                worker_ref.stopped.disconnect(self._on_stopped)
            except TypeError:
                pass
            worker_ref.stop()
            self._set_status_ui("⏸ 正在暂停...", C_YELLOW, False, "导出暂停中")
            QApplication.processEvents()
            worker_ref.wait(15000)
            # 手动清理（_on_stopped 被断开，需要自己处理）
            self.worker = None
            self.btn_start.setText("▶ 启动")
            self.btn_start.setStyleSheet(_btn_primary())
            self.btn_start.setEnabled(True)
            self.btn_login_row.setVisible(False)
            self._in_login_wait = False

        # ── 导出（参照 v42.1 格式）──
        self._set_status_ui("📊 导出中...", "#409EFF", False, "导出中")
        QApplication.processEvents()

        wb = Workbook()
        header_fill = PatternFill(start_color="07C160", end_color="07C160", fill_type="solid")
        header_font_w = Font(bold=True, size=11, color="FFFFFF")

        records = load_replied(self.cfg.get("name", "账号1"))

        # ── Sheet 1: 私信记录（v42.1 7列格式）──
        ws1 = wb.active
        ws1.title = "私信回复记录"
        ws1.append(["序号", "陌生人昵称", "联系时间", "对方消息", "我方回复", "对方后续回复", "用户手机号码"])
        for col in range(1, 8):
            cell = ws1.cell(row=1, column=col)
            cell.font = header_font_w
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        pm_history = records.get("pm_records", [])
        if pm_history:
            for i, r in enumerate(pm_history, 1):
                ws1.append([
                    i,
                    r.get("nickname", ""),
                    r.get("contact_time", r.get("time", "")),
                    r.get("first_msg", ""),
                    r.get("reply_text", ""),
                    r.get("follow_up", ""),
                    r.get("phone", "")
                ])
        else:
            ws1.append(["", "", "", "暂无记录", "", "", ""])
        ws1.column_dimensions["A"].width = 8
        ws1.column_dimensions["B"].width = 16
        ws1.column_dimensions["C"].width = 20
        ws1.column_dimensions["D"].width = 40
        ws1.column_dimensions["E"].width = 45
        ws1.column_dimensions["F"].width = 35
        ws1.column_dimensions["G"].width = 18

        # ── Sheet 2: 评论回复记录 ──
        ws2 = wb.create_sheet("评论回复记录")
        ws2.append(["序号", "回复时间", "评论昵称", "回复内容"])
        for col in range(1, 5):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font_w
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        cmt_history = records.get("cmt_records", [])
        if cmt_history:
            for i, r in enumerate(cmt_history, 1):
                ws2.append([i, r.get("time", ""), r.get("nickname", ""), r.get("reply_text", "")])
        else:
            ws2.append(["", "", "暂无记录", ""])
        ws2.column_dimensions["A"].width = 8
        ws2.column_dimensions["B"].width = 20
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 50

        wb.save(f)

        # ── 恢复运行 ──
        if was_running:
            self._set_status_ui("✅ 导出完成，恢复中...", C_ACCENT, True, "恢复运行中")
            QTimer.singleShot(300, self._toggle)

        QMessageBox.information(self, "完成",
            f"已导出至:\n{f}\n\n工作簿包含 2 张表：\n"
            f"  ① 私信回复记录（7列）\n  ② 评论回复记录\n"
            + ("\n功能已自动恢复运行。" if was_running else ""))


# ── 主窗口 ────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(960, 720)
        self.setMinimumSize(780, 560)
        self.setStyleSheet(STYLE)

        central = QWidget()
        self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ═══════════ 左侧边栏 ═══════════
        sidebar_frame = QFrame()
        sidebar_frame.setFixedWidth(220)
        sidebar_frame.setStyleSheet(f"background:{C_SIDEBAR_BG}; border:none;")
        sbl = QVBoxLayout(sidebar_frame)
        sbl.setContentsMargins(0, 0, 0, 0)
        sbl.setSpacing(0)

        sb_title = QLabel("📋 账号列表")
        sb_title.setStyleSheet(f"color:{C_TEXT_SIDEBAR}; font-size:13px; font-weight:bold; padding:16px 16px 12px 16px;")
        sbl.addWidget(sb_title)

        div = QFrame()
        div.setFrameShape(QFrame.HLine)
        div.setStyleSheet("color:#444;margin:0 12px;")
        sbl.addWidget(div)

        self.sidebar_items_layout = QVBoxLayout()
        self.sidebar_items_layout.setSpacing(0)
        sbl.addLayout(self.sidebar_items_layout)
        sbl.addStretch()

        btn_add_sidebar = QPushButton("＋ 新增账号")
        btn_add_sidebar.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: {C_ACCENT};
                border: 1px solid {C_ACCENT}; border-radius: 6px;
                padding: 8px 16px; font-size: 13px; margin: 10px 12px;
            }}
            QPushButton:hover {{ background: {C_ACCENT}; color: white; }}
        """)
        btn_add_sidebar.clicked.connect(self._add_account)
        sbl.addWidget(btn_add_sidebar)

        ver_lbl = QLabel(f"  {VERSION}")
        ver_lbl.setStyleSheet("color:#555;font-size:11px;padding:4px 16px 8px 16px;")
        sbl.addWidget(ver_lbl)

        root.addWidget(sidebar_frame)

        # ═══════════ 右侧主区域 ═══════════
        right = QVBoxLayout()
        right.setContentsMargins(0, 0, 0, 0)
        right.setSpacing(0)

        self.stack = QStackedWidget()
        self.stack.setStyleSheet(f"background:{C_MAIN_BG};")
        right.addWidget(self.stack, 1)

        # ── 日志区域 ──
        log_frame = QFrame()
        log_frame.setStyleSheet(f"background:{C_CARD_BG};border-top:1px solid {C_BORDER};")
        lfl = QVBoxLayout(log_frame)
        lfl.setContentsMargins(16, 8, 16, 10)
        lfl.setSpacing(4)
        log_hdr = QHBoxLayout()
        log_hdr.addWidget(QLabel("📋 运行日志"))
        log_hdr.addStretch()
        btn_clear_log = QPushButton("清空")
        btn_clear_log.setStyleSheet(f"""
            QPushButton {{ background:transparent; color:{C_TEXT_SECONDARY}; border:none; font-size:12px; }}
            QPushButton:hover {{ color:{C_RED}; }}
        """)
        btn_clear_log.clicked.connect(lambda: self.log_box.clear())
        log_hdr.addWidget(btn_clear_log)
        lfl.addLayout(log_hdr)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setMaximumHeight(140)
        lfl.addWidget(self.log_box)
        right.addWidget(log_frame)

        # ── 底部按钮栏 ──
        btm = QHBoxLayout()
        btm.setContentsMargins(20, 8, 20, 10)
        btm.setSpacing(12)
        btm.addStretch()
        btn_all = QPushButton("▶ 全部启动")
        btn_all.setStyleSheet(_btn_primary())
        btn_all.clicked.connect(lambda: self._all_toggle(True))
        btm.addWidget(btn_all)
        btn_stop = QPushButton("⏹ 全部停止")
        btn_stop.setStyleSheet(_btn_danger())
        btn_stop.clicked.connect(lambda: self._all_toggle(False))
        btm.addWidget(btn_stop)
        right.addLayout(btm)

        root.addLayout(right, 1)

        self._pages = []       # AccountPage 列表
        self._sidebar_items = []  # QWidget (sidebar item) 列表
        self._load_accounts()
        if len(self._pages) == 0:
            QTimer.singleShot(300, self._show_new_account_wizard)

    # ── 侧边栏操作 ──
    def _update_sidebar_name(self, idx, name):
        if 0 <= idx < len(self._sidebar_items):
            self._sidebar_items[idx].set_name(name)

    def _update_sidebar_status(self, idx, is_running, text):
        if 0 <= idx < len(self._sidebar_items):
            self._sidebar_items[idx].set_status(is_running, text)

    def _on_sidebar_click(self, idx):
        if 0 <= idx < len(self._pages):
            # 高亮选中的侧边栏项
            for i, item in enumerate(self._sidebar_items):
                item.set_active(i == idx)
            self.stack.setCurrentIndex(idx)

    def _on_sidebar_remove(self, idx):
        self._close_account(idx)

    # ── 账号管理 ──
    def _load_accounts(self):
        cfg = load_config()
        for i, ac in enumerate(cfg.get("accounts", [])):
            self._add_page(i, ac)

    def _add_page(self, idx, ac):
        page = AccountPage(idx, ac, self)
        name = ac.get("name") or f"账号{idx+1}"
        self._pages.append(page)
        self.stack.addWidget(page)

        # 创建侧边栏项
        item = SidebarItem(idx, name)
        item.clicked.connect(self._on_sidebar_click)
        item.remove_requested.connect(self._on_sidebar_remove)
        self.sidebar_items_layout.addWidget(item)
        self._sidebar_items.append(item)

        # 默认选中第一项
        if idx == 0:
            self._on_sidebar_click(0)

    def _add_account(self):
        self._show_new_account_wizard()

    def _show_new_account_wizard(self):
        """单页新建账号向导（默认值已预填，直接点确定即可）"""
        dlg = QDialog(self)
        dlg.setWindowTitle("添加账号")
        dlg.resize(520, 420)
        dlg.setMinimumWidth(480)

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(14)

        title = QLabel("📋 新建抖音客服账号")
        title.setStyleSheet("font-size:16px; font-weight:bold; color:#1A1A1A;")
        layout.addWidget(title)

        form = QFormLayout()
        form.setSpacing(10)

        le_name = QLineEdit("我的抖音账号")
        le_name.setPlaceholderText("输入抖音昵称，用于区分不同账号")
        le_name.setMinimumHeight(36)
        form.addRow("抖音昵称：", le_name)

        te_pm = QTextEdit()
        te_pm.setPlainText(DEFAULT_PM_REPLY)
        te_pm.setMaximumHeight(80)
        te_pm.setAcceptRichText(False)
        form.addRow("私信回复话术：", te_pm)

        te_cmt = QTextEdit()
        te_cmt.setPlainText(DEFAULT_CMT_REPLY)
        te_cmt.setMaximumHeight(80)
        te_cmt.setAcceptRichText(False)
        form.addRow("评论回复话术：", te_cmt)

        layout.addLayout(form)

        tip = QLabel("💡 以上为默认话术，如需自定义可直接修改，不改直接点确定即可。")
        tip.setStyleSheet("color:#888; font-size:12px;")
        tip.setWordWrap(True)
        layout.addWidget(tip)

        layout.addStretch()

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText("确定创建")
        btns.button(QDialogButtonBox.Cancel).setText("取消")
        btns.button(QDialogButtonBox.Ok).setStyleSheet(_btn_primary())
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        layout.addWidget(btns)

        if dlg.exec_() != QDialog.Accepted:
            return

        name = le_name.text().strip()
        if not name:
            QMessageBox.warning(self, "已取消", "昵称不能为空，已取消创建。")
            return
        pm_text = te_pm.toPlainText().strip() or DEFAULT_PM_REPLY
        cmt_text = te_cmt.toPlainText().strip() or DEFAULT_CMT_REPLY

        # ── 保存 ──
        cfg = load_config()
        idx = len(cfg.get("accounts", []))
        new_ac = {
            "name": name,
            "enabled": True,
            "pm_enabled": True,
            "pm_reply": pm_text,
            "comment_enabled": True,
            "comment_reply": cmt_text,
            "chrome_profile": f"chrome_profiles/account_{idx+1}"
        }
        if "accounts" not in cfg:
            cfg["accounts"] = []
        cfg["accounts"].append(new_ac)
        save_config(cfg)
        self._add_page(idx, new_ac)

        QMessageBox.information(
            self, "创建成功",
            f"「{name}」已添加！\n\n点击「▶ 启动」并扫码登录后即可开始自动回复。"
        )

    def _close_account(self, index):
        if index < 0 or index >= len(self._pages):
            return
        page = self._pages[index]
        if page.worker and page.worker.isRunning():
            page.worker.stop()
            page.worker.wait(2000)

        cfg = load_config()
        if index < len(cfg.get("accounts", [])):
            cfg["accounts"].pop(index)
            save_config(cfg)

        # 移除侧边栏项
        if index < len(self._sidebar_items):
            item = self._sidebar_items.pop(index)
            self.sidebar_items_layout.removeWidget(item)
            item.deleteLater()

        # 移除页面
        self._pages.pop(index)
        self.stack.removeWidget(page)
        page.deleteLater()

        # 更新索引
        for i, p in enumerate(self._pages):
            p.idx = i
            p._save()
        for i, item in enumerate(self._sidebar_items):
            item.idx = i

        # 选中下一个
        if self._pages:
            self._on_sidebar_click(min(index, len(self._pages) - 1))

    def _all_toggle(self, start):
        for page in self._pages:
            running = page.worker and page.worker.isRunning()
            if start and not running:
                page._toggle()
            elif not start and running:
                page._toggle()


    def _append_log(self, name, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        if msg.startswith("[green]"):
            color = C_GREEN; text = msg[7:]
        elif msg.startswith("[yellow]"):
            color = C_YELLOW; text = msg[8:]
        elif msg.startswith("[red]"):
            color = C_RED; text = msg[5:]
        elif msg.startswith("[white]"):
            color = C_TEXT; text = msg[7:]
        else:
            color = C_TEXT; text = msg
        html = f'<span style="color:#888;">{ts}</span> <b style="color:{C_GREEN};">[{_html.escape(name)}]</b> <span style="color:{color};">{_html.escape(text)}</span>'
        self.log_box.append(html)
        self.log_box.moveCursor(QTextCursor.End)
        if self.log_box.document().blockCount() > 500:
            self.log_box.clear()
            self.log_box.append('<span style="color:#888;">[日志自动清理]</span>')


def main():
    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)
    if sys.platform == "darwin":
        app.setStyle("Fusion")
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
